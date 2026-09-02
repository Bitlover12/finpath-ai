from __future__ import annotations

import csv
import io
import json
import re
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from app.data.policies import load_policies
from app.models.spending import (
    CardBenefitBreakdown,
    CardRecommendation,
    CardType,
    CardTypePreference,
    ParsedTransaction,
    SpendingCategory,
    SpendingCutItem,
    SpendingOptimizationResponse,
    SpendingProfile,
    SpendingUploadResponse,
)
from app.models.profile import UserProfile
from app.services.analyze import analyze_profile


_CATALOG_PATH = Path(__file__).resolve().parents[1] / "data" / "card_catalog_2026.json"

DISCRETIONARY_CATEGORIES = {
    SpendingCategory.DELIVERY,
    SpendingCategory.CAFE,
    SpendingCategory.SHOPPING,
    SpendingCategory.SUBSCRIPTION,
    SpendingCategory.ENTERTAINMENT,
}

CATEGORY_KEYWORDS: list[tuple[SpendingCategory, tuple[str, ...]]] = [
    (SpendingCategory.DELIVERY, ("배달의민족", "배민", "요기요", "쿠팡이츠")),
    (SpendingCategory.CAFE, ("스타벅스", "투썸", "메가커피", "메가m", "컴포즈", "이디야", "폴바셋", "커피빈", "빽다방")),
    (SpendingCategory.CONVENIENCE, ("gs25", "cu ", "cu편의점", "세븐일레븐", "이마트24")),
    (SpendingCategory.TRANSPORT, ("티머니", "카카오t", "카카오택시", "택시", "지하철", "버스", "코레일", "srt")),
    (SpendingCategory.SUBSCRIPTION, ("넷플릭스", "youtube premium", "유튜브프리미엄", "디즈니", "멜론", "지니뮤직", "쿠팡와우", "네이버플러스")),
    (SpendingCategory.TELECOM, ("skt", "sk텔레콤", "kt", "lg유플러스", "lgu+", "리브엠", "liiv m")),
    (SpendingCategory.BEAUTY, ("올리브영", "미용실", "헤어", "네일")),
    (SpendingCategory.GROCERIES, ("이마트", "롯데마트", "홈플러스", "하나로마트", "농협하나로")),
    (SpendingCategory.FUEL, ("주유소", "gs칼텍스", "s-oil", "에쓰오일", "현대오일뱅크", "sk에너지")),
    (SpendingCategory.SHOPPING, ("무신사", "29cm", "지마켓", "g마켓", "옥션", "11번가", "ssg", "네이버쇼핑", "쿠팡")),
    (SpendingCategory.CULTURE, ("cgv", "메가박스", "롯데시네마", "인터파크티켓", "yes24")),
    (SpendingCategory.FOOD, ("맥도날드", "버거킹", "롯데리아", "맘스터치", "식당", "음식점", "restaurant", "푸드")),
    (SpendingCategory.MEDICAL, ("병원", "의원", "약국", "한의원")),
    (SpendingCategory.EDUCATION, ("학원", "교육", "어학")),
]

DATE_HEADERS = {"date", "거래일", "거래일자", "이용일", "이용일자", "승인일", "승인일자", "결제일", "일자"}
DESC_HEADERS = {"description", "merchant", "가맹점", "가맹점명", "적요", "내용", "거래내용", "이용가맹점", "상호명"}
AMOUNT_HEADERS = {"amount", "금액", "거래금액", "이용금액", "결제금액", "출금액", "사용금액"}
CATEGORY_HEADERS = {"category", "카테고리", "분류"}


def load_card_catalog() -> list[dict[str, Any]]:
    with _CATALOG_PATH.open("r", encoding="utf-8") as f:
        return json.load(f)


def _amount(categories: dict[SpendingCategory, int], *keys: SpendingCategory) -> int:
    return sum(max(0, int(categories.get(k, 0))) for k in keys)


def _monthly_cap_by_spend(total: int, tiers: list[tuple[int, int]]) -> int:
    cap = 0
    for threshold, value in tiers:
        if total >= threshold:
            cap = value
    return cap


def _calc_kb_nori2(categories: dict[SpendingCategory, int], total: int) -> tuple[int, list[CardBenefitBreakdown]]:
    if total < 200_000:
        return 0, []
    parts = [
        ("커피 10%", min(round(_amount(categories, SpendingCategory.CAFE) * 0.10), 3_000)),
        ("뷰티 5%", min(round(_amount(categories, SpendingCategory.BEAUTY) * 0.05), 2_000)),
        ("편의점 5%", min(round(_amount(categories, SpendingCategory.CONVENIENCE) * 0.05), 2_000)),
    ]
    gross = sum(v for _, v in parts)
    global_cap = _monthly_cap_by_spend(total, [(200_000, 20_000), (400_000, 30_000), (600_000, 40_000), (800_000, 50_000)])
    gross = min(gross, global_cap)
    return gross, [CardBenefitBreakdown(label=k, amount=v) for k, v in parts if v > 0]


def _calc_shinhan_pick_e(categories: dict[SpendingCategory, int], total: int) -> tuple[int, list[CardBenefitBreakdown]]:
    if total < 200_000:
        return 0, []
    parts = [
        ("커피 10% 적립", min(round(_amount(categories, SpendingCategory.CAFE) * 0.10), 3_000)),
        ("편의점 10% 적립", min(round(_amount(categories, SpendingCategory.CONVENIENCE) * 0.10), 3_000)),
    ]
    return sum(v for _, v in parts), [CardBenefitBreakdown(label=k, amount=v) for k, v in parts if v > 0]


def _calc_hyundai_d(categories: dict[SpendingCategory, int], total: int) -> tuple[int, list[CardBenefitBreakdown]]:
    if total < 400_000:
        return 0, []
    core_cap = _monthly_cap_by_spend(total, [(400_000, 10_000), (800_000, 20_000), (1_200_000, 35_000)])
    extra_cap = _monthly_cap_by_spend(total, [(400_000, 5_000), (800_000, 10_000), (1_200_000, 15_000)])
    core = min(round(_amount(categories, SpendingCategory.FOOD, SpendingCategory.DELIVERY) * 0.10), core_cap)
    extra_parts = [
        ("카페 5%", min(round(_amount(categories, SpendingCategory.CAFE) * 0.05), 5_000)),
        ("편의점 5%", min(round(_amount(categories, SpendingCategory.CONVENIENCE) * 0.05), 5_000)),
        ("대중교통 5%", min(round(_amount(categories, SpendingCategory.TRANSPORT) * 0.05), 5_000)),
    ]
    extra_raw = sum(v for _, v in extra_parts)
    extra = min(extra_raw, extra_cap)
    breakdown = [CardBenefitBreakdown(label="외식·배달 10%", amount=core)] if core else []
    if extra:
        breakdown.append(CardBenefitBreakdown(label="카페·편의점·교통 5%", amount=extra))
    return core + extra, breakdown


def _calc_hyundai_z(categories: dict[SpendingCategory, int], total: int) -> tuple[int, list[CardBenefitBreakdown]]:
    if total < 500_000:
        return 0, []
    category_caps = {
        SpendingCategory.TELECOM: 5_000,
        SpendingCategory.TRANSPORT: 5_000,
        SpendingCategory.ONLINE: 4_000,
        SpendingCategory.SHOPPING: 4_000,
        SpendingCategory.CAFE: 4_000,
        SpendingCategory.CONVENIENCE: 4_000,
        SpendingCategory.DELIVERY: 4_000,
        SpendingCategory.FOOD: 2_000,
        SpendingCategory.FUEL: 5_000,
        SpendingCategory.GROCERIES: 5_000,
    }
    raw_parts: list[tuple[str, int]] = []
    for category, cap in category_caps.items():
        value = min(round(_amount(categories, category) * 0.05), cap)
        if value:
            raw_parts.append((f"{category.value} 5%", value))
    global_cap = 25_000 if total < 1_000_000 else 50_000
    gross = min(sum(v for _, v in raw_parts), global_cap)
    # Keep breakdown informational; if global cap binds, label it explicitly.
    breakdown = [CardBenefitBreakdown(label=k, amount=v) for k, v in raw_parts]
    if sum(v for _, v in raw_parts) > gross:
        breakdown.append(CardBenefitBreakdown(label="통합한도 적용", amount=-(sum(v for _, v in raw_parts) - gross)))
    return gross, breakdown


CALCULATORS = {
    "KB_NORI2_KBPAY": _calc_kb_nori2,
    "SHINHAN_PICK_E": _calc_shinhan_pick_e,
    "HYUNDAI_D": _calc_hyundai_d,
    "HYUNDAI_Z_EVERYDAY": _calc_hyundai_z,
}


def _type_allowed(card_type: str, preference: CardTypePreference) -> bool:
    return preference == CardTypePreference.BOTH or card_type == preference.value


def recommend_cards(spending: SpendingProfile) -> tuple[list[CardRecommendation], int, int]:
    categories = {k: max(0, int(v)) for k, v in spending.categories.items()}
    total = sum(categories.values())
    catalog = load_card_catalog()

    scored: list[CardRecommendation] = []
    current_net = 0
    current_seen = False

    for card in catalog:
        calculator = CALCULATORS[card["calculator"]]
        gross, breakdown = calculator(categories, total)
        eligible = total >= int(card["minimum_monthly_spend"])
        monthly_fee = round(int(card["annual_fee"]) / 12)
        net = max(0, gross - monthly_fee) if eligible else 0
        if spending.current_card_id == card["id"]:
            current_net = net
            current_seen = True

        if not _type_allowed(card["card_type"], spending.card_type_preference):
            continue
        # Core safety rule: never recommend spending more just to unlock a card benefit.
        if not eligible:
            continue

        scored.append(
            CardRecommendation(
                card_id=card["id"],
                issuer=card["issuer"],
                name=card["name"],
                card_type=CardType(card["card_type"]),
                annual_fee=int(card["annual_fee"]),
                minimum_monthly_spend=int(card["minimum_monthly_spend"]),
                eligible_without_extra_spend=True,
                estimated_gross_monthly_benefit=gross,
                estimated_monthly_fee=monthly_fee,
                estimated_net_monthly_benefit=net,
                incremental_monthly_benefit_vs_current=0,
                benefit_breakdown=breakdown,
                source_url=card["source_url"],
                checked_at=card["checked_at"],
                notes=card.get("notes", []),
            )
        )

    scored.sort(key=lambda x: (x.estimated_net_monthly_benefit, -x.annual_fee), reverse=True)
    for rec in scored:
        rec.incremental_monthly_benefit_vs_current = max(0, rec.estimated_net_monthly_benefit - current_net)

    if spending.current_card_id and not current_seen:
        current_net = 0
    best_incremental = scored[0].incremental_monthly_benefit_vs_current if scored else 0
    return scored[:3], current_net, best_incremental


def optimize_spending(profile: UserProfile, spending: SpendingProfile) -> SpendingOptimizationResponse:
    categories = {k: max(0, int(v)) for k, v in spending.categories.items()}
    total = sum(categories.values())
    discretionary_total = sum(categories.get(k, 0) for k in DISCRETIONARY_CATEGORIES)

    cut_items: list[SpendingCutItem] = []
    cut_total = 0
    for category in sorted(DISCRETIONARY_CATEGORIES, key=lambda x: x.value):
        amount = int(categories.get(category, 0))
        if amount <= 0 or spending.cut_percent <= 0:
            continue
        saving = round(amount * spending.cut_percent / 100)
        if saving <= 0:
            continue
        cut_items.append(
            SpendingCutItem(
                category=category,
                current_monthly_amount=amount,
                assumed_cut_percent=spending.cut_percent,
                monthly_saving=saving,
            )
        )
        cut_total += saving

    recommendations, current_net, card_incremental = recommend_cards(spending)
    extra = cut_total + card_incremental
    enhanced_capacity = profile.monthly_saving_capacity + extra
    enhanced_profile = profile.model_copy(update={"monthly_saving_capacity": enhanced_capacity})
    enhanced_analysis = analyze_profile(enhanced_profile, load_policies())

    return SpendingOptimizationResponse(
        total_monthly_spend=total,
        discretionary_monthly_spend=discretionary_total,
        cut_scenario_monthly_saving=cut_total,
        cut_items=cut_items,
        current_card_estimated_net_benefit=current_net,
        best_card_incremental_monthly_benefit=card_incremental,
        total_extra_monthly_saving=extra,
        enhanced_monthly_saving_capacity=enhanced_capacity,
        recommendations=recommendations,
        enhanced_analysis=enhanced_analysis,
        notice=(
            "소비 절감액은 사용자가 선택한 절감률을 적용한 시나리오이며 실제 절감 가능액을 보장하지 않습니다. "
            "카드 혜택은 현재 소비를 늘리지 않고도 실적조건을 충족하는 카드만 비교하며, 공식 상품설명서의 세부 제외조건에 따라 실제 혜택은 달라질 수 있습니다."
        ),
    )


def classify_merchant(description: str) -> SpendingCategory:
    text = f" {description.strip().lower()} "
    for category, keywords in CATEGORY_KEYWORDS:
        if any(keyword.lower() in text for keyword in keywords):
            return category
    return SpendingCategory.OTHER


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or "").strip().lower())


def _find_column(headers: list[Any], candidates: set[str]) -> int | None:
    normalized_candidates = {_normalize_header(x) for x in candidates}
    for i, header in enumerate(headers):
        if _normalize_header(header) in normalized_candidates:
            return i
    return None


def _parse_amount(value: Any) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return abs(round(value))
    text = str(value).replace(",", "").replace("원", "").replace("₩", "").strip()
    text = re.sub(r"[^0-9.\-]", "", text)
    if not text:
        return 0
    return abs(round(float(text)))


def _parse_date(value: Any) -> tuple[str | None, str | None]:
    if value is None or value == "":
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), value.strftime("%Y-%m")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%Y%m%d", "%m/%d/%Y"):
        try:
            d = datetime.strptime(text[:10], fmt)
            return d.date().isoformat(), d.strftime("%Y-%m")
        except ValueError:
            pass
    match = re.search(r"(20\d{2})[./-]?(\d{1,2})", text)
    if match:
        year, month = int(match.group(1)), int(match.group(2))
        if 1 <= month <= 12:
            return text, f"{year:04d}-{month:02d}"
    return text, None


def _category_from_value(value: Any, description: str) -> SpendingCategory:
    if value is not None:
        normalized = str(value).strip().upper()
        aliases = {
            "식비": "FOOD", "배달": "DELIVERY", "카페": "CAFE", "교통": "TRANSPORT",
            "편의점": "CONVENIENCE", "쇼핑": "SHOPPING", "구독": "SUBSCRIPTION",
            "통신": "TELECOM", "뷰티": "BEAUTY", "문화": "CULTURE", "온라인": "ONLINE",
            "마트": "GROCERIES", "주유": "FUEL", "주거": "HOUSING", "의료": "MEDICAL",
            "교육": "EDUCATION", "기타": "OTHER",
        }
        normalized = aliases.get(normalized, normalized)
        try:
            return SpendingCategory(normalized)
        except ValueError:
            pass
    return classify_merchant(description)


def _rows_to_response(rows: Iterable[list[Any]]) -> SpendingUploadResponse:
    rows = list(rows)
    if not rows:
        return SpendingUploadResponse(transactions=[], monthly_categories={}, months_count=1, total_rows=0, notice="거래내역이 비어 있습니다.")

    headers = rows[0]
    date_idx = _find_column(headers, DATE_HEADERS)
    desc_idx = _find_column(headers, DESC_HEADERS)
    amount_idx = _find_column(headers, AMOUNT_HEADERS)
    category_idx = _find_column(headers, CATEGORY_HEADERS)
    if desc_idx is None or amount_idx is None:
        raise ValueError("거래내역에는 가맹점/내용 열과 금액 열이 필요합니다.")

    transactions: list[ParsedTransaction] = []
    totals: dict[SpendingCategory, int] = defaultdict(int)
    months: set[str] = set()

    for row in rows[1:]:
        if not row or max(desc_idx, amount_idx) >= len(row):
            continue
        desc = str(row[desc_idx] or "").strip()
        amount = _parse_amount(row[amount_idx])
        if not desc or amount <= 0:
            continue
        date_value = row[date_idx] if date_idx is not None and date_idx < len(row) else None
        date_text, month = _parse_date(date_value)
        if month:
            months.add(month)
        category_value = row[category_idx] if category_idx is not None and category_idx < len(row) else None
        category = _category_from_value(category_value, desc)
        transactions.append(ParsedTransaction(date=date_text, description=desc, amount=amount, category=category))
        totals[category] += amount

    month_count = max(1, len(months))
    monthly = {category: round(amount / month_count) for category, amount in totals.items()}
    return SpendingUploadResponse(
        transactions=transactions[:500],
        monthly_categories=monthly,
        months_count=month_count,
        total_rows=len(transactions),
        notice=f"{month_count}개월 기준 월평균으로 환산했습니다. 자동분류 결과는 추천 전에 직접 수정할 수 있습니다.",
    )


def parse_spending_file(filename: str, content: bytes) -> SpendingUploadResponse:
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        text = content.decode("utf-8-sig", errors="replace")
        rows = list(csv.reader(io.StringIO(text)))
        return _rows_to_response(rows)
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        return _rows_to_response(rows)
    raise ValueError("CSV 또는 XLSX 파일만 업로드할 수 있습니다.")
